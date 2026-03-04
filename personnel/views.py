from random import choices
from django.shortcuts import render, redirect, get_object_or_404
import os
from django.conf import settings
from django.http import HttpResponse, Http404
from collections import defaultdict
from .models import *
from orders.models import *
from .forms import *
from general.decorators import *
from general.functions import *
from django.contrib import messages
from datetime import datetime
from django.utils.dateparse import parse_date
from django.views.decorators.clickjacking import xframe_options_exempt
from django.contrib.auth.decorators import login_required
from django.urls import reverse_lazy
from django.forms.models import model_to_dict
from django.template.loader import get_template, render_to_string
from xhtml2pdf import pisa
from weasyprint import HTML
import tempfile
from ajax_datatable.views import AjaxDatatableView
from bootstrap_modal_forms.generic import (
  BSModalCreateView,
  BSModalUpdateView,
  BSModalReadView,
  BSModalDeleteView
)
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from django.db import transaction, IntegrityError
from django.db.models.functions import ExtractYear, Now

# Personnel
@login_required(login_url='login')
def upload(request):
  if request.method == "POST":
    excel_file = request.FILES['upfile']
    wb = openpyxl.load_workbook(excel_file)
    sheets = ['Officer', 'EP', 'TAS', 'CHR']
    colcount = [22, 22, 22, 15]
    colreqs = [
      [0, 1, 3, 5, 6, 7, 8, 17, 21],
			[0, 1, 3, 5, 6, 7, 8, 17, 21],
			[0, 1, 3, 5, 6, 7, 8, 17, 21],
			[0, 1, 3, 5, 6, 7, 8],
    ]
    errors = []
    excel = {}
    
    for y, sheet in enumerate(sheets):
      ws = wb[sheet]
      if sheet not in excel:
        excel[sheet] = []
      for x, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        msg = []
        is_empty = False
        for i in range(colcount[y]):
          if sheet == 'CHR' and i == 1:
            continue
          else:
            if i in colreqs[y] and not row[i]:
              is_empty = True
              break
        if is_empty:
          msg.append('Missing required fields')
        if msg:
          errors.append({'sheet': sheet, 'x': x + 1, 'msg': ' | '.join(msg)})
        else:
          excel[sheet].append(row)
    if errors:
      for row in errors:
        sheet = wb[row['sheet']]
        for i in range(65, 65+colcount[sheets.index(row['sheet'])]):
          sheet[f"{chr(i)}{row['x']+1}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
        sheet[f"{chr(65+colcount[sheets.index(row['sheet'])])}{row['x']+1}"] = row['msg']
      response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
      filename = f'errors_personnel_template-{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
      response['Content-Disposition'] = 'attachment; filename=' + filename
      wb.save(response)
      messages.error(request, 'There are problems encountered during the process. Please see errors on the donwloaded file.')
      return response
    else:
      for sheetname, sheet in excel.items():
        profile = 'Officer' if sheetname == 'TAS' else sheetname
        for row in sheet:
          personnel_data = {
            'profile': profile,
            'rank': Rank.objects.get(rank=row[0]),
            'first_name': row[1].upper() if profile == 'Officer' else row[1],
            'middle_name': (row[2].upper() if profile == 'Officer' else row[2]) if row[2] else None,
            'last_name': row[3].upper() if profile == 'Officer' else row[3],
            'ext_name': (row[4].upper() if profile == 'Officer' else row[4]) if row[4] else None,
            'afpsn': None if profile == 'CHR' else row[5].upper(),
            'item_no': row[5].upper() if profile == 'CHR' else None,
            'bos': BOS.objects.get(bos=row[6]),
            'sex': 'M' if row[7] == 'M' else 'F',
            'dob': parse_date(row[7]) if profile == 'CHR' else parse_date(row[8]),
            'office': request.user.users_profile.office,
            'position': Position.objects.get(position=row[8]) if profile == 'CHR' else None,
            'desig': row[9] if row[9] else None,
            'status_assignment': SOA.objects.get(soa='Duty'),
            'hcc': row[11 if profile == 'CHR' else 19] if row[11 if profile == 'CHR' else 19] else None,
            'fos': None if profile == 'CHR' else (row[11] if row[11] else None),
            'daghq': (parse_date(row[12]) if row[12] else None) if profile == 'CHR' else None,
            'dapda': (parse_date(row[13]) if row[13] else None) if profile == 'CHR' else None,
            'rrfcd': None if profile == 'CHR' else (row[14] if row[14] else None),
            'skills': None if profile == 'CHR' else (row[15] if row[15] else None),
            'dot': (parse_date(row[16]) if row[16] else None) if profile == 'CHR' else None,
            'doc': parse_date(row[17]) if profile == 'CHR' else None,
            'dolp': (parse_date(row[18]) if row[18] else None) if profile == 'CHR' else None,            
            'soc': None if profile == 'CHR' else (row[20] if row[20] else None),
            'doret': parse_date(row[21]) if profile == 'CHR' else None,
            'auth': (row[12] if row[12] else None) if profile == 'CHR' else None,
            'act': (row[13] if row[13] else None) if profile == 'CHR' else None,
            'cti': (row[14] if row[14] else None) if profile == 'CHR' else None,
            'is_active': True,
          }
          if profile == 'CHR':
            check, created = Personnel.objects.update_or_create(item_no=personnel_data['item_no'], defaults=personnel_data)
          else:
            check, created = Personnel.objects.update_or_create(afpsn=personnel_data['afpsn'], defaults=personnel_data)
          if check:
            check.save(user=request.user)
      messages.success(request, 'Data successfully uploaded.')
    return redirect('personnel')

@login_required(login_url='login')
def download_template(request):
  file_path = os.path.join(settings.MEDIA_ROOT, 'excel_template', 'personnel_template.xlsx')
  if os.path.exists(file_path):
    with open(file_path, 'rb') as file_handle:
      response = HttpResponse(file_handle.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
      response['X-Sendfile'] = file_path
      response['Content-Length'] = os.stat(file_path).st_size
      response['Content-Disposition'] = 'attachment; filename=' + os.path.basename('uploading_template.xlsx')
      return response
  raise Http404("File not found")

def required_field(ws, cell_range, validation_type="text", formula=None, can_add_list=False):
  """
  Reusable utility to apply mandatory field constraints.
  """
  # 1. Determine the logic based on type
  if validation_type == "list":
    # Dropdown logic
    dv = DataValidation(type="list", formula1=formula, allow_blank=False, prompt="Required field. Please enter a value.")
  elif validation_type == "date":
    # Date logic: Ensure date is entered
    dv = DataValidation(type="date", operator="greaterThan", formula1='DATE(1900,1,1)', allow_blank=False, prompt="Required field. Please enter a valid date (e.g., 1995-01-23).")
  else:
    # Generic text/required logic: Ensure length > 0
    # We use the top-left cell of the range as the relative reference
    base_cell = cell_range.split(":")[0] 
    dv = DataValidation(type="custom", formula1=f"=LEN({base_cell})>0", allow_blank=False, prompt="Please select a category from the list.")

  # 2. Set common properties
  dv.showErrorMessage = True
  if validation_type == "date":
    dv.errorStyle = 'stop'
    dv.errorTitle = 'Invalid Date'
    dv.error = 'Please enter a valid date (e.g., 1995-01-23).'
  elif validation_type == "list":
    dv.errorStyle = 'warning' if can_add_list else 'stop'
    dv.errorTitle = 'Invalid Selection'
    dv.error = 'Please select a category from the list.'
  else:
    dv.errorStyle = 'stop'
    dv.errorTitle = 'Required Field'
    dv.error = 'This field is mandatory and cannot be left blank.'

  # 3. Add to sheet and link to range
  ws.add_data_validation(dv)
  dv.add(cell_range)
  
  if validation_type == "date":
    # Optional: Add date format for better UX
    for row in ws[cell_range]:
      for cell in row:
        cell.number_format = 'YYYY-MM-DD'

def optional_date(ws, cell_range):
  """
  Reusable utility to apply optional date field constraints.
  """
  dv = DataValidation(type="date", operator="greaterThan", formula1='DATE(1900,1,1)', allow_blank=True, prompt="Optional date field. Please enter a valid date (e.g., 1995-01-23).")

  # 2. Set common properties
  dv.showErrorMessage = True
  dv.errorStyle = 'stop'
  dv.errorTitle = 'Invalid Date'
  dv.error = 'Please enter a valid date (e.g., 1995-01-23).'

  # 3. Add to sheet and link to range
  ws.add_data_validation(dv)
  dv.add(cell_range)
  
  for row in ws[cell_range]:
    for cell in row:
      cell.number_format = 'YYYY-MM-DD'

def download_excel_with_dropdown(request):
  wb = openpyxl.Workbook()
  ws1 = wb.active
  ws1.title = "Officer"
  ws2 = wb.create_sheet(title="EP", index=1)
  ws3 = wb.create_sheet(title="TAS", index=2)
  ws4 = wb.create_sheet(title="CHR", index=3)
  source_ws = wb.create_sheet("SourceData")
  source_ws.sheet_state = 'hidden'
  
  len_row = 1000
  
  # required fields but not of date type
  for col in ['B', 'D', 'F']:
    required_field(ws1, f"{col}2:{col}{len_row}", validation_type="text")
    required_field(ws2, f"{col}2:{col}{len_row}", validation_type="text")
    required_field(ws3, f"{col}2:{col}{len_row}", validation_type="text")
    required_field(ws4, f"{col}2:{col}{len_row}", validation_type="text")
  
  # required dates
  for col in ['I', 'S', 'W']:
    required_field(ws1, f"{col}2:{col}{len_row}", validation_type="date")
    required_field(ws2, f"{col}2:{col}{len_row}", validation_type="date")
    required_field(ws3, f"{col}2:{col}{len_row}", validation_type="date")
  required_field(ws4, f"H2:H{len_row}", validation_type="date")
  
  # optional dates
  for col in ['N', 'O', 'R', 'T']:
    optional_date(ws1, f"{col}2:{col}{len_row}")
    optional_date(ws2, f"{col}2:{col}{len_row}")
    optional_date(ws3, f"{col}2:{col}{len_row}")
  
  # Officer Ranks
  choices_officer_ranks = Rank.objects.filter(type='Officer').values_list('rank', flat=True)
  source_ws.cell(row=1, column=1, value='Officer Ranks')
  for i, choice in enumerate(choices_officer_ranks, start=2):
    source_ws.cell(row=i, column=1, value=choice)
  formula = f"SourceData!${get_column_letter(1)}$2:${get_column_letter(1)}${len(choices_officer_ranks)+1}"
  required_field(ws1, f"A2:A{len_row}", validation_type="list", formula=formula, can_add_list=True)
  required_field(ws3, f"A2:A{len_row}", validation_type="list", formula=formula, can_add_list=True)
  
  # EP Ranks
  choices_ep_ranks = Rank.objects.filter(type='EP').values_list('rank', flat=True)
  source_ws.cell(row=1, column=2, value='EP Ranks')
  for i, choice in enumerate(choices_ep_ranks, start=2):
    source_ws.cell(row=i, column=2, value=choice)
  formula = f"SourceData!${get_column_letter(2)}$2:${get_column_letter(2)}${len(choices_ep_ranks)+1}"
  required_field(ws2, f"A2:A{len_row}", validation_type="list", formula=formula, can_add_list=True)
  
  # CHR Ranks
  choices_chr_ranks = Rank.objects.filter(type='CHR').values_list('rank', flat=True)
  source_ws.cell(row=1, column=3, value='CHR Ranks')
  for i, choice in enumerate(choices_chr_ranks, start=2):
    source_ws.cell(row=i, column=3, value=choice)
  formula = f"SourceData!${get_column_letter(3)}$2:${get_column_letter(3)}${len(choices_chr_ranks)+1}"
  required_field(ws4, f"A2:A{len_row}", validation_type="list", formula=formula, can_add_list=True)
  
  # BOS
  choices_bos = BOS.objects.exclude(ms='TAS').values_list('bos', flat=True)
  source_ws.cell(row=1, column=4, value='BOS')
  for i, choice in enumerate(choices_bos, start=2):
    source_ws.cell(row=i, column=4, value=choice)
  formula = f"SourceData!${get_column_letter(4)}$2:${get_column_letter(4)}${len(choices_bos)+1}"
  required_field(ws1, f"G2:G{len_row}", validation_type="list", formula=formula, can_add_list=True)
  required_field(ws2, f"G2:G{len_row}", validation_type="list", formula=formula, can_add_list=True)
  
  # TAS
  choices_tas = BOS.objects.filter(ms='TAS').values_list('bos', flat=True)
  source_ws.cell(row=1, column=5, value='TAS')
  for i, choice in enumerate(choices_tas, start=2):
    source_ws.cell(row=i, column=5, value=choice)
  formula = f"SourceData!${get_column_letter(5)}$2:${get_column_letter(5)}${len(choices_tas)+1}"
  required_field(ws3, f"G2:G{len_row}", validation_type="list", formula=formula, can_add_list=True)
  
  # SEX
  choices_sex = list(['M', 'F'])
  source_ws.cell(row=1, column=6, value='SEX')
  for i, choice in enumerate(choices_sex, start=2):
    source_ws.cell(row=i, column=6, value=choice)
  formula = f"SourceData!${get_column_letter(6)}$2:${get_column_letter(6)}${len(choices_sex)+1}"
  required_field(ws1, f"H2:H{len_row}", validation_type="list", formula=formula)
  required_field(ws2, f"H2:H{len_row}", validation_type="list", formula=formula)
  required_field(ws3, f"H2:H{len_row}", validation_type="list", formula=formula)
  required_field(ws4, f"G2:G{len_row}", validation_type="list", formula=formula)
    
  # SOA
  choices_soa = SOA.objects.all().values_list('soa', flat=True)
  source_ws.cell(row=1, column=7, value='SOA')
  for i, choice in enumerate(choices_soa, start=2):
    source_ws.cell(row=i, column=7, value=choice)
  formula = f"SourceData!${get_column_letter(7)}$2:${get_column_letter(7)}${len(choices_soa)+1}"
  required_field(ws1, f"K2:K{len_row}", validation_type="list", formula=formula)
  required_field(ws2, f"K2:K{len_row}", validation_type="list", formula=formula)
  required_field(ws3, f"K2:K{len_row}", validation_type="list", formula=formula)
  required_field(ws4, f"K2:K{len_row}", validation_type="list", formula=formula)
    
  # Temporary Units
  choices_temp_units = Office.objects.all().values_list('office', flat=True)
  source_ws.cell(row=1, column=8, value='Temporary Units')
  for i, choice in enumerate(choices_temp_units, start=2):
    source_ws.cell(row=i, column=8, value=choice)
  formula = f"SourceData!${get_column_letter(8)}$2:${get_column_letter(8)}${len(choices_temp_units)+1}"
  required_field(ws1, f"L2:L{len_row}", validation_type="list", formula=formula, can_add_list=True)
  required_field(ws2, f"L2:L{len_row}", validation_type="list", formula=formula, can_add_list=True)
  required_field(ws3, f"L2:L{len_row}", validation_type="list", formula=formula, can_add_list=True)
  required_field(ws4, f"L2:L{len_row}", validation_type="list", formula=formula, can_add_list=True)
    
  # CHR Plantilla Titles
  choices_positions = Position.objects.all().values_list('position', flat=True)
  source_ws.cell(row=1, column=9, value='CHR Plantilla Titles')
  for i, choice in enumerate(choices_positions, start=2):
    source_ws.cell(row=i, column=9, value=choice)
  formula = f"SourceData!${get_column_letter(9)}$2:${get_column_letter(9)}${len(choices_positions)+1}"
  required_field(ws4, f"I2:I{len_row}", validation_type="list", formula=formula, can_add_list=True)
  
  # Add Headers for Military Personnel Template
  headers = [
    {'name': 'Rank*', 'required': True, 'font_color': 'FF0000', 'width': 8},
    {'name': 'First Name*', 'required': True, 'font_color': 'FF0000', 'width': 20},
    {'name': 'Middle Name', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'Last Name*', 'required': True, 'font_color': 'FF0000', 'width': 20},
    {'name': 'Suffix', 'required': False, 'font_color': '000000', 'width': 10},
    {'name': 'AFPSN*', 'required': True, 'font_color': 'FF0000', 'width': 12},
    {'name': 'BOS*', 'required': True, 'font_color': 'FF0000', 'width': 11},
    {'name': 'Sex*', 'required': True, 'font_color': 'FF0000', 'width': 7},
    {'name': 'DOB*', 'required': True, 'font_color': 'FF0000', 'width': 16},
    {'name': 'Designation', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'SOA*', 'required': True, 'font_color': 'FF0000', 'width': 16},
    {'name': 'Temporary Unit', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'AFPOS', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'DAGHQ', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'DAPDA', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'RRFCD', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'Skills', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'DOT', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'DOC*', 'required': True, 'font_color': 'FF0000', 'width': 16},
    {'name': 'DOLP', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'HCC', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'SOC', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'DORET*', 'required': True, 'font_color': 'FF0000', 'width': 16},
  ]  
  header_style(ws1, headers)
  header_style(ws2, headers)
  header_style(ws3, headers)
  
  # Add Headers for CHR Template
  headers = [
    {'name': 'Rank*', 'required': True, 'font_color': 'FF0000', 'width': 8},
    {'name': 'First Name*', 'required': True, 'font_color': 'FF0000', 'width': 20},
    {'name': 'Middle Name', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'Last Name*', 'required': True, 'font_color': 'FF0000', 'width': 20},
    {'name': 'Suffix', 'required': False, 'font_color': '000000', 'width': 10},
    {'name': 'Plantilla Item No*', 'required': True, 'font_color': 'FF0000', 'width': 20},
    {'name': 'Sex*', 'required': True, 'font_color': 'FF0000', 'width': 7},
    {'name': 'DOB*', 'required': True, 'font_color': 'FF0000', 'width': 16},
    {'name': 'Plantilla Title*', 'required': True, 'font_color': 'FF0000', 'width': 26},
    {'name': 'Designation', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'SOA*', 'required': True, 'font_color': 'FF0000', 'width': 16},
    {'name': 'Temporary Unit', 'required': False, 'font_color': '000000', 'width': 20},
    {'name': 'HCC', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'AUTH', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'ACT', 'required': False, 'font_color': '000000', 'width': 16},
    {'name': 'CTI', 'required': False, 'font_color': '000000', 'width': 16},
  ]  
  header_style(ws4, headers)

  # 5. Prepare the response
  response = HttpResponse(
      content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
  )
  response['Content-Disposition'] = 'attachment; filename="template.xlsx"'
  
  wb.save(response)
  return response

def header_style(ws, headers = []):
  wrap_alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
  thin = Side(border_style="thin", color="000000")
  full_border = Border(top=thin, left=thin, right=thin, bottom=thin)
  
  for col_num, header_title in enumerate(headers, start=1):
    ws.column_dimensions[get_column_letter(col_num)].width = header_title['width']
    cell = ws.cell(row=1, column=col_num, value=header_title['name'])
    header_font = Font(name='Arial', bold=True, size=11, color=header_title['font_color'])
    cell.font = header_font
    cell.alignment = wrap_alignment
    cell.border = full_border

def is_ajax(meta):
  if 'HTTP_X_REQUESTED_WITH' not in meta:
    return False
  if meta['HTTP_X_REQUESTED_WITH'] == 'XMLHttpRequest':
    return True
  return False

@login_required(login_url='login')
def admin_upload(request):
  template = 'admin/admin-upload.html'
  form = AdminUploadFileForm()
  if request.method == 'POST':
    form = AdminUploadFileForm(request.POST, request.FILES)
    if form.is_valid():
      office = form.cleaned_data['office']
      excel_file = request.FILES['file']
      if not excel_file:
        messages.error(request, "Please select a file.")
        return redirect('admin-upload')

      if not excel_file.name.endswith(('.xlsx', '.xls')):
        messages.error(request, "Invalid file format. Please upload an Excel file.")
        return redirect('admin-upload')
      try:
        wb = openpyxl.load_workbook(excel_file)
        sheets = ['Officer', 'EP', 'TAS', 'CHR']
        colcount = [23, 23, 23, 16]
        colreqs = [
          [0, 1, 3, 5, 6, 7, 8, 10, 18, 22],
          [0, 1, 3, 5, 6, 7, 8, 10, 18, 22],
          [0, 1, 3, 5, 6, 7, 8, 10, 18, 22],
          [0, 1, 3, 5, 6, 7, 8, 10],
        ]
        errors = []
        excel = {}
        
        for y, sheet in enumerate(sheets):
          ws = wb[sheet]
          if sheet not in excel:
            excel[sheet] = []
          for x, row in enumerate(ws.iter_rows(min_row=2, values_only=False)):
            msg = []
            is_empty = False
            err_col = None
            if all(cell.value is None for cell in row):
              break  # Skip this row and move to the next sheet
            for i in range(colcount[y]):
              if sheet == 'CHR' and i == 1:
                continue
              else:
                if i in colreqs[y] and not row[i]:
                  is_empty = True
                  err_col = i
                  break
            if is_empty:
              msg.append(f'Missing required field in column {chr(65+err_col)}')
            if msg:
              errors.append({'sheet': sheet, 'x': x + 1, 'msg': ' | '.join(msg)})
            else:
              excel[sheet].append(row)
        if errors:
          for row in errors:
            sheet = wb[row['sheet']]
            for i in range(65, 65+colcount[sheets.index(row['sheet'])]):
              sheet[f"{chr(i)}{row['x']+1}"].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type = "solid")
            sheet[f"{chr(65+colcount[sheets.index(row['sheet'])])}{row['x']+1}"] = row['msg']
          response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
          filename = f'errors_personnel_template-{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
          response['Content-Disposition'] = 'attachment; filename=' + filename
          wb.save(response)
          messages.error(request, 'There are problems encountered during the process. Please see errors on the donwloaded file.')
          return response
        else:
          with transaction.atomic():
            rows_processed = 0
            rank_cache = {d.rank.lower(): d for d in Rank.objects.all()}
            bos_cache = {d.bos.lower(): d for d in BOS.objects.all()}
            position_cache = {d.position.lower(): d for d in Position.objects.all()}
            soa_cache = {d.soa.lower(): d for d in SOA.objects.all()}
            ds_unit_cache = {d.office.lower(): d for d in Office.objects.all()}
            office_category_cache = {d.category.lower(): d for d in OfficeCategory.objects.all()}
            for sheetname, sheet in excel.items():
              profile = 'Officer' if sheetname == 'TAS' else sheetname
              for row_num, row in enumerate(sheet, start=2):
                try:
                  # check existence of rank in cache, if not create and add to cache, else use existing
                  rank_lower = row[0].value.strip().lower()
                  if rank_lower in rank_cache:
                    rank_obj = rank_cache[rank_lower]
                  else:
                    rank_obj = Rank.objects.create(rank=row[0].value.strip(), grade=0, type=profile if profile in ['Officer', 'EP'] else 'CHR')
                    rank_cache[rank_lower] = rank_obj
                    
                  # check existence of bos in cache, if not create and add to cache, else use existing
                  bos_lower = row[6].value.strip().lower()
                  if bos_lower in bos_cache:
                    bos_obj = bos_cache[bos_lower]
                  else:
                    bos_obj = BOS.objects.create(bos=row[6].value.strip(), ms='TAS' if profile == 'TAS' else (profile if profile in ['Officer', 'EP'] else 'CHR'))
                    bos_cache[bos_lower] = bos_obj
                    
                  # check existence of position in cache, if not create and add to cache, else use existing
                  if profile == 'CHR':
                    position_lower = row[8].value.strip().lower()
                    if position_lower and position_lower in position_cache:
                      position_obj = position_cache[position_lower]
                    else:
                      position_obj = Position.objects.create(position=row[8].value.strip())
                      position_cache[position_lower] = position_obj
                  else:
                    position_obj = None
                  
                  # check existence of soa in cache, if not create and add to cache, else use existing
                  soa_lower = row[10].value.strip().lower()
                  if soa_lower in soa_cache:
                    soa_obj = soa_cache[soa_lower]
                  else:
                    soa_obj = soa_cache['duty']  # default to 'Duty' if not found
                    
                  # check existence of ds_unit in cache, if not create and add to cache, else use existing
                  if row[11].value:
                    ds_unit_lower = row[11].value.strip().lower()
                    if ds_unit_lower in ds_unit_cache:
                      ds_unit_obj = ds_unit_cache[ds_unit_lower]
                    else:
                      ds_unit_obj = Office.objects.create(office=row[11].value.strip(), office_category=office_category_cache['MS'], program=1)
                      ds_unit_cache[ds_unit_lower] = ds_unit_obj
                  else:
                    ds_unit_obj = None
                  
                  personnel_data = {
                    'profile': profile,
                    'rank': rank_obj,
                    'first_name': row[1].value.strip().upper() if profile == 'Officer' else row[1].value.strip(),
                    'middle_name': (row[2].value.strip().upper() if profile == 'Officer' else row[2].value.strip()) if row[2].value else None,
                    'last_name': row[3].value.strip().upper() if profile == 'Officer' else row[3].value.strip(),
                    'ext_name': (row[4].value.strip().upper() if profile == 'Officer' else row[4].value.strip()) if row[4].value else None,
                    'afpsn': None if profile == 'CHR' else row[5].value.strip().upper(),
                    'item_no': row[5].value.strip().upper() if profile == 'CHR' else None,
                    'bos': bos_obj,
                    'sex': 'M' if row[6].value.strip().upper() == 'M' else 'F',
                    'dob': row[7].value.date() if profile == 'CHR' else row[8].value.date(),
                    'office': Office.objects.get(office=office),
                    'position': position_obj,
                    'desig': row[9].value.strip() if row[9].value else None,
                    'status_assignment': soa_obj,
                    'ds_unit': ds_unit_obj,
                    'hcc': row[12 if profile == 'CHR' else 20].value.strip() if row[12 if profile == 'CHR' else 20].value else None,
                    'fos': None if profile == 'CHR' else (row[12].value.strip() if row[12].value else None),
                    'daghq': None if profile == 'CHR' else (row[13].value.date() if row[13].value else None),
                    'dapda': None if profile == 'CHR' else (row[14].value.date() if row[14].value else None),
                    'rrfcd': None if profile == 'CHR' else (row[15].value.strip() if row[15].value else None),
                    'skills': None if profile == 'CHR' else (row[16].value.strip() if row[16].value else None),
                    'dot': None if profile == 'CHR' else (row[17].value.date() if row[17].value else None),
                    'doc': None if profile == 'CHR' else row[18].value.date(),
                    'dolp': None if profile == 'CHR' else (row[19].value.date() if row[19].value else None),
                    'soc': None if profile == 'CHR' else (row[21].value.strip() if row[21].value else None),
                    'doret': None if profile == 'CHR' else row[22].value.date(),
                    'auth': (row[13].value.strip() if row[13].value else None) if profile == 'CHR' else None,
                    'act': (row[14].value.strip() if row[14].value else None) if profile == 'CHR' else None,
                    'cti': (row[15].value.strip() if row[15].value else None) if profile == 'CHR' else None,
                    'is_active': True,
                  }
                  
                  if profile == 'CHR':
                    check, created = Personnel.objects.update_or_create(item_no=personnel_data['item_no'], defaults=personnel_data)
                  else:
                    check, created = Personnel.objects.update_or_create(afpsn=personnel_data['afpsn'], defaults=personnel_data)
                  if check:
                    check.save(user=request.user)
                  
                  if (soa_obj.with_unit == True and soa_obj.is_school == False and soa_obj.is_medical == False and soa_obj.is_outside == False) and not ds_unit_obj:
                    raise Exception(f"Row {row_num} requires a Temporary Unit since the SOA '{soa_obj.soa}' is linked to unit assignment.")
                  if soa_obj.soa.lower() == 'a/u (local schooling)':
                    if ds_unit_obj:
                      sub = Patient.objects.filter(personnel=check).first()
                      if sub:
                        sub.delete()
                      student = Student.objects.update_or_create(personnel=check, defaults={'office': ds_unit_obj})
                    else:
                      raise Exception(f"Row {row_num} requires a Temporary Unit since the SOA '{soa_obj.soa}' is linked to local schooling assignment.")
                  if soa_obj.soa.lower() == 'a/u (hospitalized)':
                    if ds_unit_obj:
                      sub = Student.objects.filter(personnel=check).first()
                      if sub:
                        sub.delete()
                      patient = Patient.objects.update_or_create(personnel=check, defaults={'office': ds_unit_obj})
                    else:
                      raise Exception(f"Row {row_num} requires a Temporary Unit since the SOA '{soa_obj.soa}' is linked to hospitalized assignment.")
                  rows_processed += 1
                except (ValueError, TypeError, IndexError) as e:
                  # This catches errors specific to ONE row (e.g. bad formatting)
                  # We raise an exception to trigger the outer transaction rollback
                  raise Exception(f"Row {row_num} has invalid data: {e}")
            messages.success(request, 'Data successfully uploaded.')
            return redirect('admin-upload')
      except Exception as e:
        # This catches file corruption, column mismatches, or our manual raises
        messages.error(request, f"Upload failed: {str(e)}")
        return redirect('admin-upload')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'title' : 'Admin Upload',
    'uri' : 'admin-upload',
    'options': get_settings(),
    'form': form,
  }
  return render(request, template, context)

@login_required(login_url='login')
def personnel_view(request):
  template = 'personnel.html'
  unit = request.user.users_profile.office
  subunits = get_units(request.user.users_profile.office, False)
  legends = {
    'AFPSN'		      :	'AFP Serial Nr',
		'AFPOS/MOS/AFS' :	'AFP Occupational Specialty/Mode of Specialization/Air Force Specialization',
		'DAGHQ'		      :	'Date Assigned at GHQ',
		'DAPDA'		      :	'Date Assigned Present Duty Assignment',
		'DOC'		        :	'Date of Commission',
		'DOT'		        :	'Date of Termination',
		'DOB'		        :	'Date of Birth',
		'DOLP'		      :	'Date of Last Promotion',
		'HCC'		        :	'Highest Career Course',
		'RRFCD'		      :	'Rank Requirement for Current Designation',
		'SOA'		        :	'Status of Assignment',
		'SOC'		        :	'Source of Commission',
		'DORET'		      :	'Date of Retirement',
		'COS'		        :	'Contract of Service',
		'CTI'		        :	'Co-terminous with the incumbent',
  }
  context = {
    'title' : 'Personnel',
    'uri' : 'personnel',
    'options': get_settings(),
    'unit': unit,
    'subunits': subunits,
    'legends': legends,
  }
  return render(request, template, context)

class OrganicAjaxDatatableView(AjaxDatatableView):
  model = Personnel
  title = 'Personnel'
  
  initial_order = [ ["updated_at", "desc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'rank', 'foreign_field': 'rank__rank', 'title': 'Rank', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'first_name', 'title': 'First Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'middle_name', 'title': 'MI', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'last_name', 'title': 'Last Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'ext_name', 'title': 'Ext', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'afpsn', 'title': 'AFPSN', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'bos', 'foreign_field': 'bos__bos', 'title': 'BOS', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'profile', 'title': 'Profile', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'office', 'foreign_field': 'office__office', 'title': 'Unit Assignment', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', }, 
    {'name': 'status_assignment', 'foreign_field': 'status_assignment__soa', 'title': 'SOA', 'visible': True, 'className': 'text-center', },
    {'name': 'ds_unit', 'foreign_field': 'ds_unit__office', 'title': 'Temporary Unit', 'visible': True, 'className': 'text-center', },
    {'name': 'updated_at', 'title': 'Date', 'visible': True, 'className': 'text-center', },
    # {'name': 'age', 'title': 'Age', 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },    
  ]
  
  def customize_row(self, row, obj):
    # row['age'] = obj.length_service if obj.length_service is not None else None
    row['updated_at'] = obj.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    row['middle_name'] = ''.join(name[0].upper() for name in obj.middle_name.split()) if obj.middle_name else ''
    if obj.profile == 'Officer':
      row['first_name'] = obj.first_name.upper()
      row['last_name'] = obj.last_name.upper()
      row['ext_name'] = obj.ext_name.upper() if obj.ext_name else ''
    if obj.get_orders().count() == 0:
      row['tools'] = '<div class="btn-group"><a type="button" class="update-data btn btn-xs btn-primary" href="#" data-form-url="update/%s/"><i class="fas fa-edit"></i></a><a type="button" class="delete-data btn btn-xs btn-danger" href="#" data-form-url="delete/%s/"><i class="fas fa-trash"></i></a></div>' % (row['pk'], row['pk'])
    else:
      row['tools'] = '<div class="btn-group"><a type="button" class="update-data btn btn-xs btn-primary" href="#" data-form-url="update/%s/"><i class="fas fa-edit"></i></a><a type="button" class="delete-data btn btn-xs btn-danger" href="#" data-form-url="delete/%s/"><i class="fas fa-trash"></i></a><a type="button" class="view-orders btn btn-xs btn-default" href="#" data-form-url="view-orders/%s/"><i class="fas fa-file-pdf"></i></a></div>' % (row['pk'], row['pk'], row['pk'])
    return

  def get_initial_queryset(self, request=None):
    qs = super().get_initial_queryset(request)
    qs = qs.filter(is_active=True)
    office = request.user.users_profile.office    
    if request.user.groups.filter(name='Admin').exists() or request.user.groups.filter(name='Approver').exists():
      qs = qs.filter(office=office)
    return qs

@login_required(login_url='login')
def view_personnel(request, id):
  template = 'personnel/view_personnel.html'
  data = Personnel.objects.get(id=id)
  context = {
    'options': get_settings(),
    'data' : data,
  }
  return render(request, template, context)

@login_required(login_url='login')
def create_personnel(request):
  template = 'personnel/form_personnel.html'
  form = PersonnelForm()
  if request.method == 'POST':
    form = PersonnelForm(request.POST)
    if form.is_valid():
      if not is_ajax(request.META):
        f = form.save(commit=False)
        f.save(user=request.user)
      messages.success(request, 'Personnel details was successfully added!')
      return redirect('personnel')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'options': get_settings(),
    'form' : form,
  }
  return render(request, template, context)

@login_required(login_url='login')
def update_personnel(request, id):
  template = 'personnel/form_personnel.html'
  personnel = Personnel.objects.get(id=id)
  soa1 = personnel.status_assignment.soa
  form = PersonnelForm(instance=personnel)
  if request.method == 'POST':
    form = PersonnelForm(request.POST, instance=personnel)
    if form.is_valid():
      f = form.save(commit=False)
      soa2 = f.status_assignment.soa
      f.save(user=request.user)
      if soa1 != soa2:
        if soa1 == 'A/U (Local Schooling)':
          sub = Student.objects.get(personnel=personnel)
          if sub:
            sub.delete()
        if soa1 == 'A/U (Hospitalized)':
          sub = Patient.objects.get(personnel=personnel)
          if sub:
            sub.delete()
        if soa2 == 'A/U (Local Schooling)':
          if f.ds_unit:
            sub = Student(personnel=personnel, office=f.ds_unit)
          else:
            sub = Student(personnel=personnel, office=request.user.users_profile.office)          
          sub.save(user=request.user)
        if soa2 == 'A/U (Hospitalized)':
          if f.ds_unit:
            sub = Patient(personnel=personnel, office=f.ds_unit)
          else:
            sub = Patient(personnel=personnel, office=request.user.users_profile.office)          
          sub.save(user=request.user)
      messages.success(request, 'Personnel details was successfully updated!')
      return redirect('personnel')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'options': get_settings(),
    'form' : form,
  }
  return render(request, template, context)

@login_required(login_url='login')
def view_orders(request, id):
  template = 'personnel/view_orders.html'
  op = OrderPersonnel.objects.filter(personnel_id=id).values('order')
  orders = Order.objects.filter(id__in=op)
  context = {
    'options': get_settings(),
    'orders': orders,
  }
  return render(request, template, context)

class PersonnelDeleteView(BSModalDeleteView):
  model = Personnel
  template_name = 'personnel/delete_personnel.html'
  success_message = 'Personnel was successfully deleted.'
  success_url = reverse_lazy('personnel')

class SubunitAjaxDatatableView(AjaxDatatableView):
  model = Personnel
  title = 'Personnel'
  
  initial_order = [ ["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'rank', 'foreign_field': 'rank__rank', 'title': 'Rank', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'first_name', 'title': 'First Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'middle_name', 'title': 'MI', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'last_name', 'title': 'Last Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'ext_name', 'title': 'Ext', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'afpsn', 'title': 'AFPSN', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'bos', 'foreign_field': 'bos__bos', 'title': 'BOS', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'profile', 'title': 'Profile', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'office', 'foreign_field': 'office__office', 'title': 'Unit Assignment', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', }, 
    {'name': 'status_assignment', 'foreign_field': 'status_assignment__soa', 'title': 'SOA', 'visible': True, 'className': 'text-center', },
    {'name': 'ds_unit', 'foreign_field': 'ds_unit__office', 'title': 'Temporary Unit', 'visible': True, 'className': 'text-center', },
    {'name': 'updated_at', 'title': 'Date', 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },    
  ]
  
  def customize_row(self, row, obj):
    row['updated_at'] = obj.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    row['middle_name'] = ''.join(name[0].upper() for name in obj.middle_name.split()) if obj.middle_name else ''
    if obj.profile == 'Officer':
      row['first_name'] = obj.first_name.upper()
      row['last_name'] = obj.last_name.upper()
      row['ext_name'] = obj.ext_name.upper() if obj.ext_name else ''
    row['tools'] = '<div class="btn-group"><a type="button" class="view-data btn btn-xs btn-default" href="#" data-form-url="view/%s/"><i class="fas fa-eye"></i></a></div>' % (row['pk'], )
    return

  def get_initial_queryset(self, request=None):
    qs = super().get_initial_queryset(request)
    qs = qs.filter(is_active=True)
    office = request.user.users_profile.office
    subunits = get_units(office, False)
    qs = qs.filter(office__in=subunits)    
    return qs
  
class DSAjaxDatatableView(AjaxDatatableView):
  model = Personnel
  title = 'Personnel'
  
  initial_order = [ ["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'rank', 'foreign_field': 'rank__rank', 'title': 'Rank', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'first_name', 'title': 'First Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'middle_name', 'title': 'MI', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'last_name', 'title': 'Last Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'ext_name', 'title': 'Ext', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'afpsn', 'title': 'AFPSN', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'bos', 'foreign_field': 'bos__bos', 'title': 'BOS', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'profile', 'title': 'Profile', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'office', 'foreign_field': 'office__office', 'title': 'Unit Assignment', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', }, 
    {'name': 'status_assignment', 'foreign_field': 'status_assignment__soa', 'title': 'SOA', 'visible': True, 'className': 'text-center', },
    {'name': 'ds_unit', 'foreign_field': 'ds_unit__office', 'title': 'Temporary Unit', 'visible': True, 'className': 'text-center', },
    {'name': 'updated_at', 'title': 'Date', 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },
    
  ]
  
  def customize_row(self, row, obj):
    row['updated_at'] = obj.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    row['middle_name'] = ''.join(name[0].upper() for name in obj.middle_name.split()) if obj.middle_name else ''
    if obj.profile == 'Officer':
      row['first_name'] = obj.first_name.upper()
      row['last_name'] = obj.last_name.upper()
      row['ext_name'] = obj.ext_name.upper() if obj.ext_name else ''
    row['tools'] = '<div class="btn-group"><a type="button" class="view-data btn btn-xs btn-default" href="#" data-form-url="view/%s/"><i class="fas fa-eye"></i></a><a type="button" class="delete-ds btn btn-xs btn-danger" href="#" data-form-url="delete-ds/%s/"><i class="fas fa-trash"></i></a></div>' % (row['pk'], row['pk'])
    return

  def get_initial_queryset(self, request=None):
    qs = super().get_initial_queryset(request)
    qs = qs.filter(is_active=True)
    office = request.user.users_profile.office
    qs = qs.filter(status_assignment__with_unit=True, status_assignment__is_school=False, status_assignment__is_medical=False, status_assignment__is_outside=False, ds_unit=office).exclude(office=office)    
    return qs

@login_required(login_url='login')
def create_ds(request):
  template = 'personnel/form_au.html'
  form = DSForm()
  title = 'DS Personnel'
  if request.method == 'POST':
    form = DSForm(request.POST)
    if form.is_valid():
      if not is_ajax(request.META):
        pers = form.cleaned_data.get('personnel')
        personnel = Personnel.objects.get(id=pers.id)
        if personnel.status_assignment.soa == 'A/U (Local Schooling)':
          sub = Student.objects.get(personnel=personnel)
          if sub:
            sub.delete()
        if personnel.status_assignment.soa == 'A/U (Hospitalized)':
          sub = Patient.objects.get(personnel=personnel)
          if sub:
            sub.delete()
        personnel.status_assignment = SOA.objects.get(soa='DS')
        personnel.ds_unit = request.user.users_profile.office
        personnel.save(user=request.user)
      messages.success(request, 'DS personnel details was successfully added!')
      return redirect('personnel')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'options': get_settings(),
    'form' : form,
    'type' : title,
  }
  return render(request, template, context)

@login_required(login_url='login')
def delete_ds(request, id):
  template = 'personnel/delete_au.html'
  title = 'DS Personnel'
  data = get_object_or_404(Personnel, id = id)

  if request.method =="POST":
    if not is_ajax(request.META):
      data.status_assignment = SOA.objects.get(soa='Duty')
      data.ds_unit = None
      data.save(user=request.user)
    messages.success(request, 'Student was successfully deleted.')
    return redirect('personnel')
  context = {
    'options': get_settings(),
    'data' : data,
    'type' : title,
  }
  return render(request, template, context)

class StudentsAjaxDatatableView(AjaxDatatableView):
  model = Student
  title = 'Student'
  
  initial_order = [ ["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'rank', 'foreign_field': 'personnel__rank__rank', 'title': 'Rank', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'first_name', 'foreign_field': 'personnel__first_name', 'title': 'First Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'middle_name', 'foreign_field': 'personnel__middle_name', 'title': 'MI', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'last_name', 'foreign_field': 'personnel__last_name', 'title': 'Last Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'ext_name', 'foreign_field': 'personnel__ext_name', 'title': 'Ext', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'afpsn', 'foreign_field': 'personnel__afpsn', 'title': 'AFPSN', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'bos', 'foreign_field': 'personnel__bos__bos', 'title': 'BOS', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'profile', 'foreign_field': 'personnel__profile', 'title': 'Profile', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'unit', 'foreign_field': 'personnel__office__office', 'title': 'Unit Assignment', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', }, 
    {'name': 'status_assignment', 'foreign_field': 'personnel__status_assignment__soa', 'title': 'SOA', 'visible': True, 'className': 'text-center', },
    {'name': 'office', 'foreign_field': 'office__office', 'title': 'School', 'visible': True, 'className': 'text-center', },
    {'name': 'updated_at', 'title': 'Date', 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },
    
  ]
  
  def customize_row(self, row, obj):
    row['updated_at'] = obj.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    row['middle_name'] = ''.join(name[0].upper() for name in obj.personnel.middle_name.split()) if obj.personnel.middle_name else ''
    if obj.personnel.profile == 'Officer':
      row['first_name'] = obj.personnel.first_name.upper()
      row['last_name'] = obj.personnel.last_name.upper()
      row['ext_name'] = obj.personnel.ext_name.upper() if obj.personnel.ext_name else ''
    row['tools'] = '<div class="btn-group"><a type="button" class="view-data btn btn-xs btn-default" href="#" data-form-url="view/%s/"><i class="fas fa-eye"></i></a><a type="button" class="delete-student btn btn-xs btn-danger" href="#" data-form-url="delete-student/%s/"><i class="fas fa-trash"></i></a></div>' % (row['pk'], row['pk'])
    return

  def get_initial_queryset(self, request=None):
    qs = super().get_initial_queryset(request)
    office = request.user.users_profile.office
    qs = qs.filter(personnel__is_active=True, office=office)
    return qs

@login_required(login_url='login')
def create_student(request):
  template = 'personnel/form_au.html'
  form = StudentForm()
  title = 'Student'
  if request.method == 'POST':
    form = StudentForm(request.POST)
    if form.is_valid():
      if not is_ajax(request.META):
        f = form.save(commit=False)
        f.office = request.user.users_profile.office
        f.save(user=request.user)
        
        personnel = Personnel.objects.get(id=f.personnel.id)
        if personnel.status_assignment.soa == 'A/U (Hospitalized)':
          sub = Patient.objects.get(personnel=personnel)
          if sub:
            sub.delete()
        personnel.status_assignment = SOA.objects.get(soa='A/U (Local Schooling)')
        personnel.ds_unit = request.user.users_profile.office
        personnel.save(user=request.user)
      messages.success(request, 'Student details was successfully added!')
      return redirect('personnel')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'options': get_settings(),
    'form' : form,
    'type' : title,
  }
  return render(request, template, context)

@login_required(login_url='login')
def delete_student(request, id):
  template = 'personnel/delete_au.html'
  title = 'Student'
  data = get_object_or_404(Student, id = id)

  if request.method =="POST":
    if not is_ajax(request.META):
      personnel = Personnel.objects.get(id=data.personnel.id)
      personnel.status_assignment = SOA.objects.get(soa='Duty')
      personnel.ds_unit = None
      personnel.save(user=request.user)
      data.delete()
    messages.success(request, 'Student was successfully deleted.')
    return redirect('personnel')
  context = {
    'options': get_settings(),
    'data' : data,
    'type' : title,
  }
  return render(request, template, context)
  
class PatientsAjaxDatatableView(AjaxDatatableView):
  model = Patient
  title = 'Patient'
  
  initial_order = [ ["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'rank', 'foreign_field': 'personnel__rank__rank', 'title': 'Rank', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'first_name', 'foreign_field': 'personnel__first_name', 'title': 'First Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'middle_name', 'foreign_field': 'personnel__middle_name', 'title': 'MI', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'last_name', 'foreign_field': 'personnel__last_name', 'title': 'Last Name', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-left', },
    {'name': 'ext_name', 'foreign_field': 'personnel__ext_name', 'title': 'Ext', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'afpsn', 'foreign_field': 'personnel__afpsn', 'title': 'AFPSN', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'bos', 'foreign_field': 'personnel__bos__bos', 'title': 'BOS', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'profile', 'foreign_field': 'personnel__profile', 'title': 'Profile', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'unit', 'foreign_field': 'personnel__office__office', 'title': 'Unit Assignment', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', }, 
    {'name': 'status_assignment', 'foreign_field': 'personnel__status_assignment__soa', 'title': 'SOA', 'visible': True, 'className': 'text-center', },
    {'name': 'office', 'foreign_field': 'office__office', 'title': 'School', 'visible': True, 'className': 'text-center', },
    {'name': 'updated_at', 'title': 'Date', 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },
    
  ]
  
  def customize_row(self, row, obj):
    row['updated_at'] = obj.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    row['middle_name'] = ''.join(name[0].upper() for name in obj.personnel.middle_name.split()) if obj.personnel.middle_name else ''
    if obj.personnel.profile == 'Officer':
      row['first_name'] = obj.personnel.first_name.upper()
      row['last_name'] = obj.personnel.last_name.upper()
      row['ext_name'] = obj.personnel.ext_name.upper() if obj.personnel.ext_name else ''
    row['tools'] = '<div class="btn-group"><a type="button" class="view-data btn btn-xs btn-default" href="#" data-form-url="view/%s/"><i class="fas fa-eye"></i></a><a type="button" class="delete-patient btn btn-xs btn-danger" href="#" data-form-url="delete-patient/%s/"><i class="fas fa-trash"></i></a></div>' % (row['pk'], row['pk'])
    return

  def get_initial_queryset(self, request=None):
    qs = super().get_initial_queryset(request)
    office = request.user.users_profile.office
    qs = qs.filter(personnel__is_active=True, office=office)    
    return qs

@login_required(login_url='login')
def create_patient(request):
  template = 'personnel/form_au.html'
  form = PatientForm()
  title = 'Patient'
  if request.method == 'POST':
    form = PatientForm(request.POST)
    if form.is_valid():
      if not is_ajax(request.META):
        f = form.save(commit=False)
        f.office = request.user.users_profile.office
        f.save(user=request.user)
        
        personnel = Personnel.objects.get(id=f.personnel.id)
        if personnel.status_assignment.soa == 'A/U (Local Schooling)':
          sub = Student.objects.get(personnel=personnel)
          if sub:
            sub.delete()
        personnel.status_assignment = SOA.objects.get(soa='A/U (Hospitalized)')
        personnel.ds_unit = request.user.users_profile.office
        personnel.save(user=request.user)
      messages.success(request, 'Patient details was successfully added!')
      return redirect('personnel')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'options': get_settings(),
    'form' : form,
    'type' : title,
  }
  return render(request, template, context)

@login_required(login_url='login')
def delete_patient(request, id):
  template = 'personnel/delete_au.html'
  title = 'Patient'
  data = get_object_or_404(Patient, id = id)

  if request.method =="POST":
    if not is_ajax(request.META):
      personnel = Personnel.objects.get(id=data.personnel.id)
      personnel.status_assignment = SOA.objects.get(soa='Duty')
      personnel.ds_unit = None
      personnel.save(user=request.user)
      data.delete()
    messages.success(request, 'Patient was successfully deleted.')
    return redirect('personnel')
  context = {
    'options': get_settings(),
    'data' : data,
    'type' : title,
  }
  return render(request, template, context)

@login_required(login_url='login')
def action_list(request):
  template = 'action_list.html'
  context = {
    'title' : 'Action List',
    'uri' : 'action_list',
    'options': get_settings(),
  }
  return render(request, template, context)

@login_required(login_url='login')
def reports_view(request):
  template = 'reports.html'
  context = {
    'title' : 'Daily Reports',
    'uri' : 'reports',
    'options': get_settings(),
  }
  return render(request, template, context)

class ReportsAjaxDatatableView(AjaxDatatableView):
  model = Report
  title = 'Report'
  
  initial_order = [ ["is_approved", "asc"], ["is_submitted", "asc"], ["updated_at", "desc"], ["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'date', 'title': 'Date', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'office', 'foreign_field': 'office__office', 'title': 'Unit', 'placeholder': True, 'searchable': True, 'orderable': True, 'className': 'text-center', },
    {'name': 'is_submitted', 'title': 'Is Submitted?', 'visible': True, 'className': 'text-center', },
    {'name': 'is_approved', 'title': 'Is Approved?', 'visible': True, 'className': 'text-center', },
    {'name': 'updated_at', 'title': 'Updated At', 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },
    
  ]
  
  def customize_row(self, row, obj):
    row['updated_at'] = obj.updated_at.strftime('%Y-%m-%d %H:%M:%S')
    row['is_submitted'] = '<i class="fas fa-check text-success"></i>' if obj.is_submitted else ''
    row['is_approved'] = '<i class="fas fa-check text-success"></i>' if obj.is_approved else ''
    if obj.is_submitted:
      if self.request.user.groups.filter(name='Admin').exists() :
        row['tools'] = '<div class="btn-group"><a href="view/%s/" type="button" class="view-data btn btn-xs btn-info"><i class="fas fa-eye"></i></a><a type="button" class="delete-data btn btn-xs btn-danger" href="#" data-form-url="delete/%s/"><i class="fas fa-trash"></i></a></div>' % (row['pk'], row['pk'])
      else:
        row['tools'] = '<div class="btn-group"><a href="view/%s/" type="button" class="view-data btn btn-xs btn-info"><i class="fas fa-eye"></i></a></div>' % (row['pk'])
    else:
      row['tools'] = '<div class="btn-group"><a href="view/%s/" type="button" class="view-data btn btn-xs btn-info"><i class="fas fa-eye"></i></a><a type="button" class="delete-data btn btn-xs btn-danger" href="#" data-form-url="delete/%s/"><i class="fas fa-trash"></i></a></div>&nbsp;<a type="button" class="submit-data btn btn-xs btn-default" href="#" data-form-url="submit/%s/"><i class="fas fa-paper-plane"></i> Submit For Approval</a>' % (row['pk'], row['pk'], row['pk'])
    return

  def get_initial_queryset(self, request=None):
    qs = super().get_initial_queryset(request)
    office = request.user.users_profile.office    
    if request.user.groups.filter(name='Admin').exists() :
      qs = qs.filter(office=office)
    elif request.user.groups.filter(name='Approver').exists() :
      qs = qs.filter(office=office, is_submitted=True)
    return qs

@login_required(login_url='login')
def create_report(request):
  template = 'reports/form_report.html'
  unit = request.user.users_profile.office
  subunits = get_units(request.user.users_profile.office, False)
  form = ReportDateFilterForm()  
  if request.method == 'POST':
    form = ReportDateFilterForm(request.POST)
    if form.is_valid():
      date = request.POST.get('date')
      is_exist = Report.objects.filter(date=parse_date(date), office=unit).exists()
      if not is_exist:
        report = Report(
          date = parse_date(date),
          office = unit,
          created_by = request.user,
        )
        report.save(user=request.user)
        messages.success(request, 'Report details was successfully added!')
      else:
        messages.error(request, 'Report already exists.')
      return redirect('reports')
    else:
      messages.error(request, 'Please correct the error.')
  context = {
    'title' : 'Report',
    'uri' : 'form_report',
    'options': get_settings(),
    'form' : form,
    'unit': unit,
    'subunits': subunits,
  }
  return render(request, template, context)

@login_required(login_url='login')
def view_report(request, id):
  template = 'reports/view_report.html'
  unit = request.user.users_profile.office
  subunits = get_units(request.user.users_profile.office, False)
  report = Report.objects.get(id=id)
  comments = ReportComment.objects.filter(report=report).order_by('created_at')
  context = {
    'title' : 'Report',
    'uri' : 'view_report',
    'options': get_settings(),
    'unit': unit,
    'subunits': subunits,
    'report': report,
    'comments': comments,
  }
  return render(request, template, context)

class ReportDeleteView(BSModalDeleteView):
  model = Report
  template_name = 'reports/delete_report.html'
  success_message = 'Report was successfully deleted.'
  success_url = reverse_lazy('reports')

@login_required(login_url='login')
def submit_report(request, id):
  report = Report.objects.get(id=id)
  if report:
    report.is_submitted = True
    report.save(user=request.user)
    messages.success(request, 'Report was successfully submitted!')
  return redirect('reports')

@login_required(login_url='login')
def approve_report(request, id):
  report = Report.objects.get(id=id)
  if report:
    types = ['ORGANIC', 'DS', 'STUDENTS', 'PATIENTS']
    for type in types:
      if type == 'ORGANIC':
        personnel = Personnel.objects.filter(is_active=True, office=report.office).order_by('office__office', '-rank__type', '-rank__grade', 'rank__id', 'last_name', 'first_name')
      elif type == 'DS':
        personnel = Personnel.objects.filter(is_active=True, status_assignment__with_unit=True, status_assignment__is_school=False, status_assignment__is_medical=False, status_assignment__is_outside=False, ds_unit=report.office).exclude(office=report.office).order_by('office__office', '-rank__type', '-rank__grade', 'rank__id', 'last_name', 'first_name')
      elif type == 'STUDENTS':
        sub = Student.objects.select_related().filter(personnel__is_active=True, office=report.office).order_by('personnel__office__office', '-personnel__rank__type', '-personnel__rank__grade', 'personnel__rank__id', 'personnel__last_name', 'personnel__first_name').values('personnel')
        personnel = Personnel.objects.filter(id__in=sub)
      elif type == 'PATIENTS':
        sub = Patient.objects.filter(personnel__is_active=True, office=report.office).order_by('personnel__office__office', '-personnel__rank__type', '-personnel__rank__grade', 'personnel__rank__id', 'personnel__last_name', 'personnel__first_name').values('personnel')
        personnel = Personnel.objects.filter(id__in=sub)
      for person in personnel:
        if type == 'ORGANIC':
          pers = PersonnelHist()
        elif type == 'DS':
          pers = PersonnelHist()
        elif type == 'STUDENTS':
          pers = StudentHist()
        elif type == 'PATIENTS':
          pers = PatientHist()
        pers.report = report
        pers.id_pers = person.id
        pers.profile = person.profile
        pers.rank = person.rank.rank
        pers.first_name = person.first_name
        pers.middle_name = person.middle_name
        pers.last_name = person.last_name
        pers.ext_name = person.ext_name
        pers.afpsn = person.afpsn
        pers.bos = person.bos.bos
        pers.office = person.office.office
        pers.fos = person.fos
        pers.daghq = person.daghq
        pers.dapda = person.dapda
        pers.status_assignment = person.status_assignment.soa
        pers.rrfcd = person.rrfcd
        pers.desig = person.desig
        pers.position = person.position.position if person.position else None
        pers.item_no = person.item_no
        pers.skills = person.skills
        pers.dot = person.dot
        pers.doc = person.doc
        pers.dob = person.dob
        pers.dolp = person.dolp
        pers.hcc = person.hcc
        pers.soc = person.soc
        pers.doret = person.doret
        pers.sex = person.sex
        pers.auth = person.auth
        pers.act = person.act
        pers.vacant = person.vacant
        pers.cti = person.cti
        pers.remarks = person.remarks
        pers.ds_unit = person.ds_unit.office if person.ds_unit else None
        pers.outside_unit = person.outside_unit.outside_unit if person.outside_unit else None
        pers.is_active = person.is_active
        pers.save(user=request.user)
    report.is_approved = True
    report.approved_by = request.user
    report.save(user=request.user)
    messages.success(request, 'Report was successfully approved!')
  return redirect('reports')

@login_required(login_url='login')
@xframe_options_exempt
def generate_report(request, type, id=None):
  date=''
  report = None
  if id:
    report = Report.objects.get(id=id)
    office = report.office
    date = report.date
  else:
    office = request.user.users_profile.office
    if request.GET.get('date'):
      date = parse_date(request.GET.get('date'))
  if (report and not report.is_approved) or not report:
    if type == 'ORGANIC':
      personnel = Personnel.objects.filter(is_active=True, office=office).order_by('office__office', '-rank__type', '-rank__grade', 'rank__id', 'last_name', 'first_name')
    elif type == 'SUBUNIT':
      subunits = get_units(office, False)
      personnel = Personnel.objects.filter(is_active=True,office__in=subunits).order_by('office__office', '-rank__type', '-rank__grade', 'rank__id', 'last_name', 'first_name')
    elif type == 'DS':
      personnel = Personnel.objects.filter(is_active=True, status_assignment__with_unit=True, status_assignment__is_school=False, status_assignment__is_medical=False, status_assignment__is_outside=False, ds_unit=office).exclude(office=office).order_by('office__office', '-rank__type', '-rank__grade', 'rank__id', 'last_name', 'first_name')
    elif type == 'STUDENTS':
      personnel = Student.objects.select_related().filter(personnel__is_active=True, office=office).order_by('personnel__office__office', '-personnel__rank__type', '-personnel__rank__grade', 'personnel__rank__id', 'personnel__last_name', 'personnel__first_name')
    elif type == 'PATIENTS':
      personnel = Patient.objects.select_related().filter(personnel__is_active=True, office=office).order_by('personnel__office__office', '-personnel__rank__type', '-personnel__rank__grade', 'personnel__rank__id', 'personnel__last_name', 'personnel__first_name')
    else:
      personnel = Personnel.objects.all(is_active=True, )
  else:
    if type == 'ORGANIC':
      personnel = PersonnelHist.objects.filter(is_active=True, office=office.office, report=report)
    elif type == 'SUBUNIT':
      print(office)
      subunits = get_units(office, False)
      print(subunits)
      reports = Report.objects.filter(office__in=subunits, date=date)
      print(reports)
      personnel = PersonnelHist.objects.filter(is_active=True, report__in=reports)
    elif type == 'DS':
      personnel = PersonnelHist.objects.filter(is_active=True, ds_unit=office.office, report=report).exclude(office=office.office)
    elif type == 'STUDENTS':
      personnel = StudentHist.objects.filter(is_active=True, report=report)
    elif type == 'PATIENTS':
      personnel = PatientHist.objects.filter(is_active=True, report=report)
    else:
      personnel = Personnel.objects.all(is_active=True, )
  
  template = 'reports/print_report.html'
  user = request.user
  context = {
    'title' : 'Report',
    'uri' : 'form_report',
    'options': get_settings(),
    'type': type,
    'personnel': personnel,
    'date': date,
    'user': user,
    'report': report,
  }
  
  html_string = render_to_string(template, context, request=request)
  html = HTML(string=html_string, base_url=request.build_absolute_uri())
  result = html.write_pdf()

  # Creating http response
  response = HttpResponse(content_type='application/pdf;')
  response['Content-Disposition'] = 'inline; filename=report.pdf'
  response['Content-Transfer-Encoding'] = 'binary'
  with tempfile.NamedTemporaryFile(delete=True) as output:
    output.write(result)
    output.flush()
    output.seek(0)
    # output = open(output.name, 'rb')
    response.write(output.read())

  return response

@login_required(login_url='login')
def add_comment(request, id_report):
  report = Report.objects.get(id=id_report)
  if request.method == 'POST':
    comment = ReportComment()
    comment.report = report
    comment.comment = request.POST.get('comment')
    comment.save(user=request.user)
    messages.success(request, 'Comment was successfully added!')
  return redirect('view-report', id=report.id)

class ReportCommentDeleteView(BSModalDeleteView):
  model = ReportComment
  template_name = 'reports/delete_comment.html'
  success_message = 'Comment was successfully deleted.'
  
  def form_valid(self, form):
    self.object = self.get_object()
    self.object.delete()
    return redirect('view-report', id=self.kwargs.get('id_report'))

# SOA
@login_required(login_url='login')
def soa(request):
  template = 'layouts/table.html'
  context = {
    'title' : 'Status of Assignment',
    'modal_size' : 'md',
    'options': get_settings(),
  }
  return render(request, template, context)

class SOAAjaxDatatableView(AjaxDatatableView):
  model = SOA
  title = 'Status of Assignment'
  
  initial_order = [["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'soa', 'title': 'SOA', 'visible': True, 'className': 'text-center', },
    {'name': 'description', 'title': 'Description', 'visible': True, 'className': 'text-center', },
    {'name': 'is_effective', 'title': 'Effectiveness', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'with_unit', 'title': 'With Carrying Unit?', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'reporting', 'title': 'Reporting', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'is_school', 'title': 'School', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'is_medical', 'title': 'Medical Facility', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'is_outside', 'title': 'Outside Organization', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },
  ]
  
  def customize_row(self, row, obj):
    row['tools'] = '<button type="button" class="update-data bs-modal btn btn-xs btn-primary" data-form-url="update/%s/"><i class="fas fa-pencil-alt"></i> Edit</button>&nbsp;<button type="button" class="delete-data bs-modal btn btn-xs btn-danger" data-form-url="delete/%s/"><i class="fas fa-trash"></i> Delete</button>' % (row['pk'], row['pk'])
    return

class SOACreateView(BSModalCreateView):
  template_name = 'layouts/form-data.html'
  form_class = SOAForm
  success_message = 'SOA was successfully created.'
  success_url = reverse_lazy('soa')
  
  def get_context_data(self, **kwargs):
    ctx = super(SOACreateView, self).get_context_data(**kwargs)
    ctx['segment'] = 'SOA'
    ctx['mode'] = 'Create'
    ctx['options'] = get_settings()
    return ctx

class SOAUpdateView(BSModalUpdateView):
  model = SOA
  template_name = 'layouts/form-data.html'
  form_class = SOAForm
  success_message = 'SOA was successfully updated.'
  success_url = reverse_lazy('soa')
  
  def get_context_data(self, **kwargs):
    ctx = super(SOAUpdateView, self).get_context_data(**kwargs)
    ctx['segment'] = 'SOA'
    ctx['mode'] = 'Update'
    ctx['options'] = get_settings()
    return ctx
  
class SOADeleteView(BSModalDeleteView):
  model = SOA
  template_name = 'layouts/form-data-delete.html'
  success_message = 'SOA was successfully deleted.'
  success_url = reverse_lazy('soa')
  
  def get_context_data(self, **kwargs):
    ctx = super(SOADeleteView, self).get_context_data(**kwargs)
    ctx['segment'] = 'SOA'
    ctx['mode'] = 'Delete'
    ctx['options'] = get_settings()
    ctx['object_name'] = self.object.soa
    return ctx

# Position
@login_required(login_url='login')
def position(request):
  template = 'layouts/table.html'
  context = {
    'title' : 'Position',
    'modal_size' : 'md',
    'options': get_settings(),
  }
  return render(request, template, context)

class PositionAjaxDatatableView(AjaxDatatableView):
  model = Position
  title = 'Position'
  
  initial_order = [["id", "asc"], ]
  length_menu = [[10, 20, 50, 100, -1], [10, 20, 50, 100, 'all']]
  search_values_separator = '+'

  column_defs = [
    {'name': 'id', 'visible': False, 'orderable': True},
    {'name': 'position', 'title': 'Position', 'visible': True, 'className': 'text-center', },
    {'name': 'description', 'title': 'Description', 'visible': True, 'className': 'text-center', },
    {'name': 'parenthetical', 'title': 'Parenthetical', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'grade', 'title': 'Grade', 'searchable': False, 'visible': True, 'className': 'text-center', },
    {'name': 'tools', 'title': 'Tools', 'placeholder': True, 'searchable': False, 'orderable': False, 'className': 'text-center', },
  ]
  
  def customize_row(self, row, obj):
    row['tools'] = '<button type="button" class="update-data bs-modal btn btn-xs btn-primary" data-form-url="update/%s/"><i class="fas fa-pencil-alt"></i> Edit</button>&nbsp;<button type="button" class="delete-data bs-modal btn btn-xs btn-danger" data-form-url="delete/%s/"><i class="fas fa-trash"></i> Delete</button>' % (row['pk'], row['pk'])
    return

class PositionCreateView(BSModalCreateView):
  template_name = 'layouts/form-data.html'
  form_class = PositionForm
  success_message = 'Position was successfully created.'
  success_url = reverse_lazy('position')
  
  def get_context_data(self, **kwargs):
    ctx = super().get_context_data(**kwargs)
    ctx['segment'] = 'Position'
    ctx['mode'] = 'Create'
    ctx['options'] = get_settings()
    return ctx

class PositionUpdateView(BSModalUpdateView):
  model = Position
  template_name = 'layouts/form-data.html'
  form_class = PositionForm
  success_message = 'Position was successfully updated.'
  success_url = reverse_lazy('position')
  
  def get_context_data(self, **kwargs):
    ctx = super(PositionUpdateView, self).get_context_data(**kwargs)
    ctx['segment'] = 'Position'
    ctx['mode'] = 'Update'
    ctx['options'] = get_settings()
    return ctx
  
class PositionDeleteView(BSModalDeleteView):
  model = Position
  template_name = 'layouts/form-data-delete.html'
  success_message = 'Position was successfully deleted.'
  success_url = reverse_lazy('position')
  
  def get_context_data(self, **kwargs):
    ctx = super(PositionDeleteView, self).get_context_data(**kwargs)
    ctx['segment'] = 'Position'
    ctx['mode'] = 'Delete'
    ctx['options'] = get_settings()
    ctx['object_name'] = self.object.position
    return ctx

@login_required(login_url='login')
def age_report_views(request):
  template = 'reports/age_report.html'
  form = AgeReportFilterForm()
  context = {
    'title' : 'Age Recap',
    'options': get_settings(),
    'form' : form,
  }
  return render(request, template, context)

@login_required(login_url='login')
@xframe_options_exempt
def generate_age_report(request):
  category = request.GET.get('category')
  office_category = OfficeCategory.objects.filter(id=category).first() if category else None
  print(f'Generating age report for category: {category}')
  age_range = list(range(17, 58))  # 17 to 57
  grand_total_per_age = {age: 0 for age in age_range}
  matrices = {'Officer': [], 'EP': []}  # Separate matrices for Officers and Enlisted Personnel
  total_per_age_types = {'Officer': {age: 0 for age in age_range}, 'EP': {age: 0 for age in age_range}}  # Separate totals for each type
  sub_total_types = {'Officer': 0, 'EP': 0}  # Subtotals for each type
  for rank_type in ['Officer', 'EP']:
    ranks = [f'{'O' if rank_type == "Officer" else "E"}-{i}' for i in range(10, 0, -1)]  # O-10 down to O-1

    # 2. Fetch data (Optimized with .only)
    if category:
      personnel_list = Personnel.objects.only('rank', 'dob').annotate(computed_age=ExtractYear(Now()) - ExtractYear('dob')).filter(computed_age__gte=17, computed_age__lte=57).filter(is_active=True, rank__type__in=[rank_type]).filter(office__office_category=category)
    else:
      personnel_list = Personnel.objects.only('rank', 'dob').annotate(computed_age=ExtractYear(Now()) - ExtractYear('dob')).filter(computed_age__gte=17, computed_age__lte=57).filter(is_active=True, rank__type__in=[rank_type])
    # return personnel_list
    # 3. Create a tally map: {(rank, age): count}
    # Using defaultdict simplifies adding to keys that don't exist yet
    tally = defaultdict(int)
    
    for p in personnel_list:
      # Accessing the @property here
      if p.age in age_range:
        tally[(f'{'O' if rank_type == "Officer" else "E"}-{p.rank.grade}', p.age)] += 1

    # 4. Build the Matrix for the template
    matrix = []
    total_per_age = {age: 0 for age in age_range}
    sub_total = 0

    for rank in ranks:
      row_data = []
      row_sum = 0
      for age in age_range:
        count = tally.get((rank, age), 0)
        # Use None or 0; we'll handle the "empty cell" look in the template
        row_data.append(count if count > 0 else "")
        
        row_sum += count
        total_per_age[age] += count
        grand_total_per_age[age] += count
        sub_total += count
          
      matrix.append({
        'rank': rank,
        'columns': row_data,
        'total': row_sum
      })
    matrices[rank_type] = matrix
    total_per_age_types[rank_type] = total_per_age.values()
    sub_total_types[rank_type] = sub_total
  template = 'reports/age_recap.html'
  user = request.user
  context = {
    'title' : 'Age Recap',
    'uri' : 'form_report',
    'options': get_settings(),
    'matrices': matrices,
    'age_range': age_range,
    'total_per_age_types': total_per_age_types,
    'sub_total_types': sub_total_types,
    'grand_total_per_age': grand_total_per_age.values(),
    'grand_total': sum(grand_total_per_age.values()),
    'report_date': timezone.now(),
    'office_category': office_category,
  }
  
  html_string = render_to_string(template, context, request=request)
  html = HTML(string=html_string, base_url=request.build_absolute_uri())
  result = html.write_pdf()

  # Creating http response
  response = HttpResponse(content_type='application/pdf;')
  response['Content-Disposition'] = 'inline; filename=age_recap.pdf'
  response['Content-Transfer-Encoding'] = 'binary'
  with tempfile.NamedTemporaryFile(delete=True) as output:
    output.write(result)
    output.flush()
    output.seek(0)
    # output = open(output.name, 'rb')
    response.write(output.read())

  return response
